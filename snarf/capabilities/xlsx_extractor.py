import io

from openpyxl import load_workbook

from snarf.capabilities.base import Capability


class XlsxExtractor(Capability):
    name = "xlsx_extractor"

    @property
    def available(self) -> bool:
        return True

    def extract_text(self, xlsx_bytes: bytes) -> str:
        workbook = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
        parts = []
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell) for cell in row if cell is not None]
                if cells:
                    parts.append(" | ".join(cells))
        return "\n".join(parts).strip()
